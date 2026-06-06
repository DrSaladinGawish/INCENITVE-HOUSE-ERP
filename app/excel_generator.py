import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)
EGP_FORMAT = '#,##0.00'


def _style_sheet(ws, headers: list[str], data_rows: list[list], freeze: bool = True):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(data_rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
    for col_idx in range(1, len(headers) + 1):
        max_len = max((len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(data_rows) + 2)), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
    if freeze:
        ws.freeze_panes = "A2"


def generate_excel(sheets: list[dict]) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for s in sheets:
        ws = wb.create_sheet(title=s.get("name", "Sheet")[:31])
        _style_sheet(ws, s.get("headers", []), s.get("rows", []))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_csv(headers: list[str], rows: list[list]) -> str:
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()
